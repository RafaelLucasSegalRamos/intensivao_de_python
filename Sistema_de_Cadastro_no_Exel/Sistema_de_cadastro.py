import customtkinter
from tkinter import *
from tkinter import messagebox
import openpyxl
import os
from openpyxl import Workbook

customtkinter.set_appearance_mode('System')


class App(customtkinter.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()

    def layout_config(self):
        self.title('Sistema de gestão de clientes')
        self.geometry('800x500')

    def appearence(self):
        self.lb_apm = customtkinter.CTkLabel(self, text="Tema", bg_color='transparent', text_color=('#000', '#fff')). \
            place(x=25, y=430)

        self.opt_apm = customtkinter.CTkOptionMenu(self, values=['System', 'Dark', 'Light'], command=self.change_apm,
                                         corner_radius=10) \
            .place(x=25, y=460)

    def todo_sistema(self):
        frame = customtkinter.CTkFrame(self, width=740, height=60, corner_radius=50, fg_color='teal').place(x=30, y=20)
        title = customtkinter.CTkLabel(frame, text='Aplicativo de gerenciamento de clientes',
                             font=('Copperplate Gothic Bold', 24), bg_color='teal', text_color='white').place(x=110,
                                                                                                              y=35)
        span = customtkinter.CTkLabel(self, text='Por favor, preencha todos os campos do formulário!',
                            font=('Copperplate Gothic Bold', 20), text_color=('Black', 'White')).place(x=40, y=90)

        ficheiro = 'Clientes.xlsx'
        if os.path.exists(ficheiro):
            pass
        else:
            ficheiro = Workbook()
            folha = ficheiro.active
            folha['A1'] = "Nome Completo"
            folha['B1'] = "Contato"
            folha['C1'] = "Idade"
            folha['D1'] = "Gênero"
            folha['E1'] = "Endereço"
            folha['F1'] = "Observação"

            ficheiro.save('Clientes.xlsx')

        def submit():

            # pegando dados dos campos
            name = namevalue.get()
            contato = contvalue.get()
            ano = agevalue.get()
            genero = genderCB.get()
            endereco = addresvalue.get()
            opcao = optentry.get(0.0, END)

            if name == '' or contato == '' or ano == '' or genero == '' or genero == '' or opcao == '':
                messagebox.showerror('Sistema', 'O campos não foram preenchidos corretamente!!')
            else:
                ficheiro = openpyxl.load_workbook('Clientes.xlsx')
                folha = ficheiro.active
                folha.cell(column=1, row=folha.max_row + 1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contato)
                folha.cell(column=3, row=folha.max_row, value=ano)
                folha.cell(column=4, row=folha.max_row, value=genero)
                folha.cell(column=5, row=folha.max_row, value=endereco)
                folha.cell(column=6, row=folha.max_row, value=opcao)

                ficheiro.save(r'Clientes.xlsx')
                messagebox.showinfo("Sistema", 'Dados salvos com sucesso!')

        def clear():
            namevalue.set('')
            contvalue.set('')
            agevalue.set('')
            addresvalue.set('')
            optentry.delete(0.0, END)

        # texto variaveis
        namevalue = StringVar()
        contvalue = StringVar()
        agevalue = StringVar()
        addresvalue = StringVar()

        nameentry = customtkinter.CTkEntry(self, placeholder_text='Nome...', textvariable=namevalue, font=('Arial', 14),
                                 width=300,
                                 fg_color='transparent')
        contentry = customtkinter.CTkEntry(self, placeholder_text='Telefone...', textvariable=contvalue, font=('Arial', 14,),
                                 width=300,
                                 fg_color='transparent')
        ageentry = customtkinter.CTkEntry(self, placeholder_text='Idade...', textvariable=agevalue, font=('Arial', 14), width=150,
                                fg_color='transparent')

        genderCB = customtkinter.CTkComboBox(self, width=150, values=['Masculino', 'Feminino', 'Outro'], font=('Arial', 14))
        genderCB.set("Masculino")

        addresentry = customtkinter.CTkEntry(self, placeholder_text='Endereço...', font=('Arial', 14), width=250,
                                   fg_color='transparent', textvariable=addresvalue)

        optentry = customtkinter.CTkTextbox(self, font=('Arial', 18), width=620, height=135, border_color='#aaa', border_width=2,
                                  fg_color='transparent')

        lbname = customtkinter.CTkLabel(self, text='Preencha com o nome do cliente:',
                              font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))
        lbcontact = customtkinter.CTkLabel(self, text='Preencha com o contato do cliente:',
                                 font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))
        lbage = customtkinter.CTkLabel(self, text='Idade do cliente:',
                             font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))
        lbgender = customtkinter.CTkLabel(self, text='Gênero do cliente:',
                                font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))
        lbaddres = customtkinter.CTkLabel(self, text='Endereço do cliente:',
                                font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))
        lbopt = customtkinter.CTkLabel(self, text='Caso tenha mais alguma observação preencha:',
                             font=('Copperplate Gothic Bold', 14), text_color=('Black', 'White'))

        btnsubmit = customtkinter.CTkButton(self, text='CADASTRAR', font=('Copperplate Gothic Bold', 18), width=80, height=40,
                                  command=submit, fg_color='#151', hover_color='#131').place(x=600, y=450)
        btnsubmit = customtkinter.CTkButton(self, text='Limpar campos'.upper(), font=('Copperplate Gothic Bold', 18), width=80,
                                  height=40,
                                  command=clear, fg_color='#555', hover_color='#333').place(x=350, y=450)
        # posicionamento dos elementos

        lbname.place(x=50, y=120)
        nameentry.place(x=50, y=150)

        lbcontact.place(x=400, y=120)
        contentry.place(x=400, y=150)

        lbage.place(x=50, y=190)
        ageentry.place(x=50, y=220)

        lbgender.place(x=250, y=190)
        genderCB.place(x=250, y=220)

        lbaddres.place(x=450, y=190)
        addresentry.place(x=450, y=220)

        lbopt.place(x=80, y=260)
        optentry.place(x=80, y=290)

    def change_apm(self, new_appearence):
        customtkinter.set_appearance_mode(new_appearence)


if __name__ == '__main__':
    app = App()
    app.mainloop()
